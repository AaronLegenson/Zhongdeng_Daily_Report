# -*- coding: utf-8 -*-

import re
import os
import json
import time
import openpyxl
import urllib3
import hashlib
import requests
import datetime
import pymysql
import pymysql.cursors
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from strings import *


def print_df(df):
    pd.set_option("display.max_rows", 500)
    pd.set_option("display.max_columns", 500)
    pd.set_option("display.width", 1000)
    print(df)


def sha256(filename):
    with open(filename, "rb") as f:
        data = f.read()
    file_md5 = hashlib.sha256(data).hexdigest()
    return file_md5


def now_time_string(string_format=STRING_FORMAT_FULL_SECOND):
    return stamp_to_string(time.time(), string_format)


def stamp_to_string(stamp, string_format=STRING_FORMAT_FULL_SECOND):
    return time.strftime(string_format, time.localtime(stamp))


def string_to_stamp(string, string_format=STRING_FORMAT_FULL_SECOND):
    return time.mktime(time.strptime(string, string_format))


def remove_brackets(string):
    """
    移除字符串中的中英文括号，主要因为excel的sheet名要避开括号
    :param string: 原字符串
    :return: 移除括号后的字符串
    """
    return string.replace("(", "").replace(")", "").replace("（", "").replace("）", "")


def build_df_business_full(df_business, df_tab_dict: dict, dic_activity, dic_activity_all):
    """
    在business表的基础上补充额外的三列
    :param df_business:
    :param df_tab_dict:
    :param dic_activity:
    :param dic_activity_all:
    :return:
    """
    df_business_full = df_business.copy(deep=True)
    # print_df(df_business_full)
    business_names = df_business_full[STRING_COL_BUSINESS_NAME]
    query_times_in_ten_days, query_times_all, string_col_query_org_count = [], [], []
    for one_business in business_names:
        # 近十天查询次数
        if dic_activity.get(one_business):
            query_times_in_ten_days.append(dic_activity.get(one_business)[2])
        else:
            query_times_in_ten_days.append(0)

        # 总查询次数
        if dic_activity_all.get(one_business):
            query_times_all.append(dic_activity_all.get(one_business)[2])
        else:
            query_times_all.append(0)

        # 总查询企业数量
        if one_business in df_tab_dict:
            string_col_query_org_count.append(len(df_tab_dict.get(one_business)))
        else:
            string_col_query_org_count.append(0)
    # print(len(query_times_in_ten_days), query_times_in_ten_days)
    # print(len(query_times_all), query_times_all)
    # print(len(string_col_query_org_count), string_col_query_org_count)
    df_business_full[STRING_COL_QUERY_TIMES_IN_TEN_DAYS] = query_times_in_ten_days
    df_business_full[STRING_COL_QUERY_TIMES_ALL] = query_times_all
    df_business_full[STRING_COL_QUERY_ORG_COUNT] = string_col_query_org_count
    df_business_full = df_business_full[[
        STRING_COL_BUSINESS_NAME,
        STRING_COL_ACCOUNT_TYPE,
        STRING_COL_ACCOUNT_COUNT,
        STRING_COL_WHETHER_FP,
        STRING_COL_RELEASE_DATE,
        STRING_COL_QUERY_TIMES_IN_TEN_DAYS,
        STRING_COL_QUERY_TIMES_ALL,
        STRING_COL_QUERY_ORG_COUNT,
        STRING_COL_TEAM,
        STRING_COL_TEAM_CONTACT
    ]]
    return df_business_full


def build_newly_added(df: pd.DataFrame, legal_list):
    previous_stamp, today_stamp = previous_period(legal_list)
    df_newly_added = pd.DataFrame(columns=df.columns.values.tolist())
    for index, row in df.iterrows():
        if previous_stamp <= string_to_stamp(row.get(STRING_COL_QUERY_TIME)[:19]) <= today_stamp:
            df_newly_added = df_newly_added.append(row)
    df[STRING_COL_KEYWORD_COUNT] = df[STRING_COL_KEYWORD_COUNT].astype(int)
    df[STRING_COL_BUSINESS_ID] = df[STRING_COL_BUSINESS_ID].astype(int)
    df[STRING_COL_USER_ID] = df[STRING_COL_USER_ID].astype(int)
    df_newly_added.reset_index(drop=True, inplace=True)
    df_newly_added.index += 1
    return df_newly_added, stamp_to_string(previous_stamp, STRING_FORMAT_FULL_MINUTE), stamp_to_string(today_stamp, STRING_FORMAT_FULL_MINUTE)


def build_tab(df, business_name, dic):
    """
    单家机构的查询结果汇总
    :param df: 全量log的表
    :param business_name: 机构名
    :param dic:
    :return:
    """
    df_tab = df[(df[STRING_COL_BUSINESS_NAME] == business_name)].copy(deep=True)
    # print_df(df_tab)
    df_tab[STRING_COL_QUERY_COUNT] = [1] * len(df_tab)
    df_group = df_tab[[
        STRING_COL_ORG_NAME, STRING_COL_KEYWORD_COUNT]].groupby(
        [STRING_COL_ORG_NAME], as_index=False).sum()
    df_group[STRING_COL_QUERY_COUNT] = df_tab[[
        STRING_COL_ORG_NAME, STRING_COL_KEYWORD_COUNT]].groupby(
        [STRING_COL_ORG_NAME], as_index=False).count()[STRING_COL_KEYWORD_COUNT]
    registration_count, file_count, page_count = [], [], []
    # print(df_group[STRING_COL_ORG_NAME])
    for one_org_name in df_group[STRING_COL_ORG_NAME]:
        # tab_detail = get_tab_detail(session, one_org_name)
        if dic.get(one_org_name):
            tab_detail = dic.get(one_org_name)
        else:
            tab_detail = [-1, -1, -1]
        registration_count.append(tab_detail[0])
        file_count.append(tab_detail[1])
        page_count.append(tab_detail[2])
    df_group[STRING_COL_REGISTRATION_COUNT] = registration_count
    df_group[STRING_COL_FILE_COUNT] = file_count
    df_group[STRING_COL_PAGE_COUNT] = page_count
    df_group = df_group[[STRING_COL_ORG_NAME, STRING_COL_QUERY_COUNT, STRING_COL_KEYWORD_COUNT, STRING_COL_REGISTRATION_COUNT, STRING_COL_FILE_COUNT, STRING_COL_PAGE_COUNT]]
    df_group.index = df_group.index + 1
    # print_df(df_group)
    return df_group


def build_instructions():
    df = pd.DataFrame()
    df["使用说明"] = [
        "1.在“开通机构信息”页面点击“开通机构”超链接转到机构查询明细界面；",
        "2.在机构查询明细页面点击“返回主页”回到“开通机构信息”页面；",
        "3.试用账户有效期默认为一个月，体验账户暂无到期日；",
        "4.全量详情为所有机构用户的查询明细。"
    ]
    return df


def build_activity(df):
    """
    :param df: 表：全量查询日志(机构用户)
    :return:
    """
    df = df[[STRING_COL_QUERY_TIME, STRING_COL_BUSINESS_NAME, STRING_COL_TEAM]]
    # print_df(df)
    today = datetime.date.fromtimestamp(time.time())  # - datetime.timedelta(days=5)
    today_stamp = time.mktime(today.timetuple())
    # df_ten_days = pd.DataFrame(columns=[STRING_COL_QUERY_TIME, STRING_COL_BUSINESS_NAME, STRING_COL_TEAM])
    dic_activity = dict()
    for index, row in df.iterrows():
        if today_stamp - 86400 * 10 <= string_to_stamp(row.get(STRING_COL_QUERY_TIME)[:19]) <= today_stamp:
            # df_ten_days = df_ten_days.append(row)
            row_business_name: str = row.get(STRING_COL_BUSINESS_NAME)
            row_team: str = row.get(STRING_COL_TEAM)
            if not dic_activity.get(row_business_name):
                dic_activity[row_business_name] = [row_business_name, row_team, 1]
            else:
                tmp = dic_activity.get(row_business_name)
                tmp[2] += 1
                dic_activity[row_business_name] = tmp

    # 20210618新增，顺便数个总数
    dic_activity_all = dict()
    for index, row in df.iterrows():
        row_business_name: str = row.get(STRING_COL_BUSINESS_NAME)
        if not dic_activity_all.get(row_business_name):
            dic_activity_all[row_business_name] = [row_business_name, None, 1]
        else:
            tmp = dic_activity_all.get(row_business_name)
            tmp[2] += 1
            dic_activity_all[row_business_name] = tmp

    # print_df(df_ten_days)
    activity_list = list(dic_activity.values())
    activity_list.sort(key=lambda x: -x[2])

    df_activity = pd.DataFrame(columns=[STRING_COL_RANK, STRING_COL_BUSINESS_NAME, STRING_COL_TEAM, STRING_COL_QUERY_TIMES_IN_TEN_DAYS])
    for i, row in enumerate(activity_list):
        df_activity = df_activity.append({STRING_COL_BUSINESS_NAME: row[0], STRING_COL_TEAM: row[1], STRING_COL_QUERY_TIMES_IN_TEN_DAYS: row[2]}, ignore_index=True)
    df_activity.index += 1
    df_activity[STRING_COL_RANK] = df_activity.index
    # print_df(df_activity)
    return df_activity, dic_activity, dic_activity_all


def set_excel_style(filename, not_empty_business_list):
    book = openpyxl.load_workbook(filename)
    book.active = 1
    # 初始化字体样式
    font_bold = Font(name='等线',
                     size=11,
                     bold=True,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='FF000000',
                     outline='None')

    font_not_bold = Font(name='等线',
                         size=11,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='FF000000',
                         outline='None')

    font_link = Font(name='等线',
                     size=11,
                     italic=False,
                     vertAlign=None,
                     underline='single',
                     strike=False,
                     color='1810d2',
                     outline='None')

    border_none = Border(top=Side(), bottom=Side(), left=Side(), right=Side())
    border_full = Border(top=Side(border_style='thin'), bottom=Side(border_style='thin'),
                         left=Side(border_style='thin'), right=Side(border_style='thin'))
    border_double = Border(top=Side(border_style='thin'), bottom=Side(border_style='double'), left=Side(), right=Side())
    for one_sheet_name in book.sheetnames:
        sheet: Worksheet = book[one_sheet_name]
        # if sheet.title in not_empty_business_list:
        #     tmp_max_column = sheet.max_column
        #     sheet.cell(1, tmp_max_column + 1).value = STRING_COL_REGISTRATION_COUNT
        #     sheet.cell(1, tmp_max_column + 2).value = STRING_COL_FILE_COUNT
        #     sheet.cell(1, tmp_max_column + 3).value = STRING_COL_PAGE_COUNT
        max_row = sheet.max_row
        max_column = sheet.max_column
        
        if sheet.title != STRING_SHEET_INSTRUCTIONS:
            # 设置整体的字体
            for row in sheet.rows:
                for cell in row:
                    cell.font = font_not_bold
                    cell.border = border_full
        
            # 加粗
            for i in range(1, sheet.max_column + 1):
                sheet.cell(1, i).font = font_bold
    
            # 居中
            for i in range(max_row):
                for j in range(max_column):
                    if i == 0:
                        sheet.cell(i + 1, j + 1).fill = PatternFill(fill_type='solid', fgColor="B4C6E7")
                    sheet.cell(i + 1, j + 1).alignment = Alignment(horizontal='center', vertical='center')
    
            # # 列宽
            # for i in range(max_column):
            #     sheet.column_dimensions[get_column_letter(i + 1)].width = 35
                
        # 使用说明Sheet
        else:
            sheet.sheet_properties.tabColor = "F75000"
            row_offset = 2
            col_offset = 1
            sheet.insert_rows(1, row_offset)
            sheet.insert_cols(1, col_offset)
            
            for row in sheet.rows:
                for cell in row:
                    cell.font = font_not_bold
                    cell.border = border_none
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            sheet.cell(1 + row_offset, 1 + col_offset).font = font_bold
            
            sheet.column_dimensions[get_column_letter(1 + col_offset)].width = 60
        
        # 超链接
        not_empty_business_list_no_brackets = [remove_brackets(item) for item in not_empty_business_list]
        # print("not_empty_business_list_no_brackets:", not_empty_business_list_no_brackets)
        if sheet.title == STRING_SHEET_BUSINESS:
            for i in range(2, max_row + 1):
                business_name = sheet.cell(i, 1).value
                if business_name in not_empty_business_list:

                    sheet.cell(i, 1).value = "=HYPERLINK(\"#{0}!A1\",\"{1}\")".format(remove_brackets(business_name), sheet.cell(i, 1).value)
                    sheet.cell(i, 1).font = font_link
                    # link = "#'%s'!A1" % (business_name)
                    # sheet.cell(i, 1).hyperlink = link
                    # print(sheet.cell(i, 1).hyperlink.target)
                    
        elif sheet.title in not_empty_business_list_no_brackets:
            # link = "#'%s'!A1" % (STRING_SHEET_BUSINESS)
            # sheet.cell(2, max_column + 2).value = STRING_EXCEL_RETURN_HOME
            # sheet.cell(2, max_column + 2).hyperlink = link
            sheet.cell(2, max_column + 2).value = "=HYPERLINK(\"#{0}!A1\",\"{1}\")".format(STRING_SHEET_BUSINESS, STRING_EXCEL_RETURN_HOME)
            sheet.cell(2, max_column + 2).font = font_link


        # 将每一列，单元格列宽最大的列宽值存到字典里，key:列的序号从1开始(与字典num_str_dic中的key对应)；value:列宽的值
        max_column_dict = {}

        # 遍历全部列
        for i in range(1, max_column + 1):
            # 遍历每一列的全部行
            for j in range(1, max_row + 1):
                column = 0
                # 获取j行i列的值
                sheet_value = sheet.cell(j, i).value
                if "!A1" in str(sheet_value):
                    sheet_value = str(sheet_value).split("\"")[3]
                # 通过列表生成式生成字符列表，将当前获取到的单元格的str值的每一个字符放在一个列表中（列表中一个元素是一个字符）
                sheet_value_list = [k for k in str(sheet_value)]
                # print(sheet_value_list, end="")
                # 遍历当前单元格的字符列表
                for v in sheet_value_list:
                    # 判定长度，一个数字或一个字母，单元格列宽+=1.1，其它+=2.2（长度可根据需要自行修改，经测试一个字母的列宽长度大概为1）
                    # if v.isdigit() or v.isalpha():
                    if ord(v) < 256:
                        column += 1.3
                    else:
                        column += 2.2
                # 当前单元格列宽与字典中的对比，大于字典中的列宽值则将字典更新。如果字典没有这个key，抛出异常并将值添加到字典中
                # print(column)
                try:
                    if not max_column_dict.get(i) or column > max_column_dict[i]:
                        max_column_dict[i] = column
                except Exception as e:
                    print("Error:", e)
                    max_column_dict[i] = column
        # 此时max_column_dict字典中已存有当前sheet的所有列的最大列宽值，直接遍历字典修改列宽
        for key, value in max_column_dict.items():
            sheet.column_dimensions[get_column_letter(key)].width = value
        
        book.save(filename)


def get_week_day(string, string_format=STRING_FORMAT_FULL_SECOND):
    """
    通过一个时间字符串得到今天星期几(1-7)
    :param string: 时间字符串
    :param string_format: 时间字符串的格式
    :return: 一个int，结果为1-7
    """
    return datetime.datetime.fromtimestamp(time.mktime(time.strptime(string, string_format))).isoweekday()


def previous_period(legal_list):
    """
    依据推送weekday的list，例如[1,3,5]，得到当日0时时间戳，和上一个发送日的0时时间戳
    :param legal_list: 推送weekday的list，例如[1,3,5]
    :return: 返回两个时间戳，当日0时时间戳，和上一个发送日的0时时间戳
    """
    if len(legal_list) == 0:
        print("Error: illegal weekday list!")
        return time.time(), time.time()
    today = datetime.date.fromtimestamp(time.time())  # - datetime.timedelta(days=5)
    today_stamp = time.mktime(today.timetuple())
    previous = today - datetime.timedelta(days=1)
    while previous.isoweekday() not in legal_list:
        previous -= datetime.timedelta(days=1)
    previous_stamp = time.mktime(previous.timetuple())
    return previous_stamp, today_stamp


def unite_html(content_list, team_name):
    """
    将html的部件拼成完整的html字符串
    :param content_list: 部件的list，每个item有3个部分：类型, 小标题, 小标题对应的html字符串
    :param team_name: 如果是team类型邮件，附上team_name
    :return: 返回完整的html字符串，可以直接用于发送，放在txt改格式html的话本地也可以打开看
    """
    content_html = STRING_MAIL_TEXT_HEAD + STRING_MAIL_TEXT_TITLE.format(team_name) + STRING_MAIL_TEXT_README
    for item in content_list:
        if item[0] == "log":
            new_content = STRING_MAIL_TEXT_PART_LOG.format(
                item[1],
                item[2].replace(STRING_MAIL_TEXT_ALIGN_RIGHT, "")
            )
            content_html += new_content
        elif item[0] == "highlight":
            new_content = STRING_MAIL_TEXT_PART_NORMAL.format(
                item[1],
                item[2].replace(STRING_MAIL_TEXT_ALIGN_RIGHT, "")
            )
            content_html += new_content.replace("class=\"dataframe\"", "class=\"dataframe\" id=\"table_highlight\"")
        elif item[0] == "none_blue":
            new_content = STRING_MAIL_TEXT_PART_NONE_BLUE.format(
                item[1],
                item[2]
            )
            content_html += new_content
        elif item[0] == "none_red":
            new_content = STRING_MAIL_TEXT_PART_NONE_RED.format(
                item[1],
                item[2]
            )
            content_html += new_content
        else:
            new_content = STRING_MAIL_TEXT_PART_NORMAL.format(
                item[1],
                item[2].replace(STRING_MAIL_TEXT_ALIGN_RIGHT, "")
            )
            if "开通机构信息" in item[1]:
                new_content = new_content.replace("正式账号", "<font color=\"#00FF00\">正式账号</font>")  # #00FF00
                new_content = new_content.replace("试用账号", "<font color=\"#00BB00\">试用账号</font>")  # #00BB00
            content_html += new_content.replace("class=\"dataframe\"", "class=\"dataframe\" id=\"table_normal\"")
    content_html += STRING_MAIL_TEXT_TAIL
    return content_html


class Log:
    def __init__(self):
        self.logs = []
        self.flag = False
        self.log = STRING_EMPTY

    def print(self, string, log_end=None):
        """
        类似普通的print，除此之外往实例里面也存放对应的内容
        :param string:
        :param log_end: 若缺失，自动换行，否则\n被该字符替代
        :return:
        """
        if self.flag:
            full_log = string
        else:
            full_log = STRING_FORMAT_3.format(now_time_string(), STRING_SPACE, string)
        if log_end is not None:
            print(full_log, end=STRING_EMPTY)
            self.logs.append(full_log)
            self.flag = True
        else:
            print(full_log)
            full_log += STRING_PLAIN_ENTER
            self.logs.append(full_log)
            self.flag = False
        self.log += full_log


class MySQLConnection:
    def __init__(self):
        self.connection_params = {
            "host": "172.16.24.74",  # "172.16.24.75",
            "port": 3306,
            "user": "sam",
            "passwd": "qxtrJnSs9*C$ZMCb",
            "db": "sam_bi_analysis"
        }
        self.connection = None  # pymysql.connections.Connection()
        self.cursor = None  # pymysql.cursors.Cursor()

    def get_user(self):
        # 用于获取数据库上的机构用户/内部用户表
        # 外部客户的
        self.cursor.execute(
            """
                select `business_id`, `business_name`, cast(`user_id` as char) as `user_id`, `user_name`, if(`roles` like "%发票验真%", "是", "-") as `fp`, `release_date`, `team`, `team_contact`, `roles`
                from sam_bi_analysis.zhongdeng_user_org_info_prd
                where `team` is not null
                and use_flag = 1
                and `roles` like "%中登网查询%"
                order by `release_date`;
            """
        )
        # df_outer = pd.DataFrame()
        df_outer = pd.DataFrame(columns=[
            STRING_COL_BUSINESS_ID,
            STRING_COL_BUSINESS_NAME,
            STRING_COL_USER_ID,
            STRING_COL_USER_NAME,
            STRING_COL_WHETHER_FP,
            STRING_COL_RELEASE_DATE,
            STRING_COL_TEAM,
            STRING_COL_TEAM_CONTACT,
            STRING_COL_ROLES])
        for item in self.cursor.fetchall():
            row = {
                STRING_COL_BUSINESS_ID: item[0],
                STRING_COL_BUSINESS_NAME: item[1],
                STRING_COL_USER_ID: item[2],
                STRING_COL_USER_NAME: item[3],
                STRING_COL_WHETHER_FP: item[4],
                STRING_COL_RELEASE_DATE: item[5],
                STRING_COL_TEAM: item[6],
                STRING_COL_TEAM_CONTACT: item[7],
                STRING_COL_ROLES: item[8]
            }
            df_outer = df_outer.append(row, ignore_index=True)
        df_outer.sort_values(by=[STRING_COL_RELEASE_DATE], ascending=[True], inplace=True)
        df_outer.reset_index(drop=True, inplace=True)
        df_outer.index += 1
        # print_df(df_outer)

        # 内部员工的
        self.cursor.execute(
            """
                select `business_id`, `business_name`, cast(`user_id` as char) as `user_id`, `user_name`, if(`roles` like "%发票验真%", "是", "-") as `fp`, `release_date`, `roles`
                from sam_bi_analysis.zhongdeng_user_org_info_prd
                where `team` is null
                and use_flag = 1
                and business_name = "天津聚量商业保理有限公司"
                and `roles` like "%中登网查询%"
                order by `release_date`;
            """
        )
        # df_inner = pd.DataFrame()
        df_inner = pd.DataFrame(columns=[
            STRING_COL_BUSINESS_ID,
            STRING_COL_BUSINESS_NAME,
            STRING_COL_USER_ID,
            STRING_COL_USER_NAME,
            STRING_COL_WHETHER_FP,
            STRING_COL_RELEASE_DATE,
            STRING_COL_ROLES])
        for item in self.cursor.fetchall():
            row = {
                STRING_COL_BUSINESS_ID: item[0],
                STRING_COL_BUSINESS_NAME: item[1],
                STRING_COL_USER_ID: item[2],
                STRING_COL_USER_NAME: item[3],
                STRING_COL_WHETHER_FP: item[4],
                STRING_COL_RELEASE_DATE: item[5],
                STRING_COL_ROLES: item[6]
            }
            df_inner = df_inner.append(row, ignore_index=True)
        try:
            df_book = self.get_oa_contact_book()
        except Exception as e:
            print("Error in get contact book from oa.fusionfintrade.com:", e)
            print("Try local contact book instead, while it's not a long-term solution!")
            df_book = pd.read_excel("normal/contact_book_20210621.xlsx")
        df_inner = df_inner.merge(df_book, on=STRING_COL_USER_NAME)
        df_inner[STRING_COL_BUSINESS_NAME] = ["内部"] * len(df_inner)
        df_inner[STRING_COL_TEAM_CONTACT] = ["(空)"] * len(df_inner)
        df_inner = df_inner[[
            STRING_COL_BUSINESS_ID,
            STRING_COL_BUSINESS_NAME,
            STRING_COL_USER_ID,
            STRING_COL_USER_NAME,
            STRING_COL_WHETHER_FP,
            STRING_COL_RELEASE_DATE,
            STRING_COL_TEAM,
            STRING_COL_TEAM_CONTACT,
            STRING_COL_ROLES
        ]]
        df_inner.sort_values(by=[STRING_COL_RELEASE_DATE], ascending=[True], inplace=True)
        df_inner.reset_index(drop=True, inplace=True)
        df_inner.index += 1
        return df_outer, df_inner

    def get_business(self):
        # 用于获取数据库上的机构信息表
        self.cursor.execute(
            """
            select
                t1.`business_name`,
                t2.`account_type`, 
                t2.`account_count`,
                t2.`fp`,
                t2.`release_date`,
                t2.`team`,
                t2.`team_contact`
            from
                sam_bi_analysis.business_info_prd as t1
                left join (
                    select
                        info.`business_name` as `business_name`,
                        if (max(info.`account_type`) = 2, "正式账号", if(max(info.`account_type`) = 1, "试用账号", "体验账号")) as `account_type`,
                        count(1) as `account_count`,
                        if (max(info.`fp_bool`) = 1, "是", "-") as `fp`,
                        min(info.`release_date`) as `release_date`,
                        max(`team`) as `team`,
                        max(`team_contact`) as `team_contact`
                    from
                    (
                    select `business_name`, `user_name`, if(`roles` like "%发票验真%", 1, 0) as `fp_bool`, `release_date`, if(`roles` like "%中登网查询账号（正式）%", 2, if(`roles` like "%中登网查询账号（试用）%", 1, 0)) as `account_type`, `team`, `team_contact`
                    from sam_bi_analysis.zhongdeng_user_org_info_prd
                    where `team` is not null
                    and `use_flag` = 1
                    and `roles` like "%中登网查询%"
                    ) as info
                    group by business_name
                    order by `release_date`
                ) as t2
                on t1.`business_name` = t2.`business_name`
            ;
            """
        )
        # df = pd.DataFrame()
        df = pd.DataFrame(columns=[
            STRING_COL_BUSINESS_NAME,
            STRING_COL_ACCOUNT_TYPE,
            STRING_COL_ACCOUNT_COUNT,
            STRING_COL_WHETHER_FP,
            STRING_COL_RELEASE_DATE,
            STRING_COL_TEAM,
            STRING_COL_TEAM_CONTACT
        ])
        for item in self.cursor.fetchall():
            row = {
                STRING_COL_BUSINESS_NAME: item[0],
                STRING_COL_ACCOUNT_TYPE: item[1],
                STRING_COL_ACCOUNT_COUNT: item[2],
                STRING_COL_WHETHER_FP: item[3],
                STRING_COL_RELEASE_DATE: item[4],
                STRING_COL_TEAM: item[5],
                STRING_COL_TEAM_CONTACT: item[6]
            }
            df = df.append(row, ignore_index=True)
        df.sort_values(by=[STRING_COL_RELEASE_DATE], ascending=[True], inplace=True)
        df.reset_index(drop=True, inplace=True)
        df.index += 1
        return df

    def get_log(self):
        self.cursor.execute(
            """
            select `enterprise_name`, `query_time`, `keyword_count`, `user_id`
            from sam_bi_analysis.user_log_history_prd
            where `enterprise_name` not like "% 等%家企业%"
            """
        )
        df = pd.DataFrame(columns=[
            STRING_COL_ORG_NAME, STRING_COL_QUERY_TIME, STRING_COL_KEYWORD_COUNT, STRING_COL_USER_ID
        ])
        for item in self.cursor.fetchall():
            row = {
                STRING_COL_ORG_NAME: item[0],
                STRING_COL_QUERY_TIME: item[1],
                STRING_COL_KEYWORD_COUNT: item[2],
                STRING_COL_USER_ID: item[3]
            }
            df = df.append(row, ignore_index=True)
        df.sort_values(by=[STRING_COL_QUERY_TIME], ascending=[False], inplace=True)
        df.reset_index(drop=True, inplace=True)
        df.index += 1
        return df

    def get_oa_contact_book(self):
        self.cursor.execute(
            """
            select `name`, `team`
            from sam_bi_analysis.oa_contact_book_prd
            """
        )
        df = pd.DataFrame(columns=[
            STRING_COL_USER_NAME,
            STRING_COL_TEAM,
        ])
        for item in self.cursor.fetchall():
            row = {
                STRING_COL_USER_NAME: item[0],
                STRING_COL_TEAM: item[1]
            }
            df = df.append(row, ignore_index=True)
        df.sort_values(by=[STRING_COL_TEAM, STRING_COL_USER_NAME], inplace=True)
        df.reset_index(drop=True, inplace=True)
        df.index += 1
        return df

    def get_enterprise_register_info(self):
        """
        获得一个字典，查询企业名对应的三列，enterprise_name->[registration_count, file_count, page_count]
        :param org_name_list:
        :return:
        """
        dic = dict()
        self.cursor.execute(
            """
            select `enterprise_name`, `registration_count`, `file_count`, `page_count`
            from sam_bi_analysis.szdw_enterprise_register_info
            """
        )
        for item in self.cursor.fetchall():
            dic[item[0]] = [item[1], item[2], item[3]]
        return dic

    def connect(self):
        self.connection = pymysql.connect(**self.connection_params)
        self.cursor = self.connection.cursor()

    def close(self):
        self.connection.commit()
        self.cursor.close()
        self.connection.close()
            

if __name__ == "__main__":
    # mc = MySQLConnection()
    # mc.connect()
    # print("hello")
    # df_res = mc.get_log()
    # print("ok")
    # mc.close()
    # print_df(df_res)
    # df_book_test = pd.read_excel("normal/contact_book.xlsx")
    # print_df(df_book_test)
    # df_business = get_business_names_new()
    # print("df_business:")
    # print_df(df_business)
    # df_user_outer, df_user_inner = get_user_new()
    # print("df_user_outer:")
    # print_df(df_user_outer)
    # df_log_all = read_log_files("logs")
    # print("df_log_all:")
    # print_df(df_log_all)
    # df_outer = df_log_all.merge(df_user_outer, on=STRING_COL_USER_ID)
    # print("df_outer:")
    # print_df(df_outer)


    # df_res = pd.read_excel("saves/SAM产品各首代团队客户使用日报-全公司_20210607_141754.xlsx", sheet_name="全量查询日志(机构用户)", engine="openpyxl")
    # build_activity(df_res)
    # file = open("logs/user_query.log", "rb")
    # for line in file:
    #     strs = line.decode()
    #     print(strs)
    # lines = [line.decode() for line in file]
    # print(lines)

    # debug_user_info_all()
    # df_res = get_user_new()
    # print_df(df_res)
    # strings = [
    #     "金融机构合作首席代表何伟强团队",
    #     "金融机构合作首席代表刘志刚团队",
    #     "金融机构合作首席代表刘明团队",
    #     "金融机构合作首席代表孟庆波团队",
    #     "金融机构合作首席代表张俊涛团队",
    #     "金融机构合作首席代表沈彦炜团队",
    #     "金融机构合作首席代表王晓光团队",
    #     "金融机构合作首席代表刘芊妤团队(筹)",
    #     "医药行业事业部"
    # ]
    # strings.sort()
    # time.sleep(5)
    # print(strings)

    # df_test = read_log_files("logs")
    # print_df(df_test)
    # print(string_to_stamp("2021-05-08 18:11:47,488"[:19], STRING_FORMAT_FULL_SECOND))
    # test = previous_period([4, 7])
    # print(stamp_to_string(test))
    # print("1".isdigit())
    # print("s".isdigit())
    # print("我".isdigit())
    # print("1".isalpha())
    # print("s".isalpha())
    # print("我".isalpha())
    # print(ord("1"))
    # print(ord("s"))
    # print(ord("我"))


    # session = requests.Session()
    # import time
    # t0 = time.time()
    # for i in range(0, 1):
    #     res = get_tab_detail_new(["杭州杭富轨道交通有限公司", "山东海王银河医药有限公司"])
    #     print(res)
    # t1 = time.time()
    # print(t1 - t0, "s")


    # set_excel_style("saves/中登网开通用户信息清单_20210428_152248.xlsx", ["远东国际融资租赁有限公司", "民生银行上海分行"])
    # res = pd.read_excel("saves/中登网开通用户信息清单_20210428_151002.xlsx", sheet_name=STRING_SHEET_DETAILS)
    # build_tab(res, "远东国际融资租赁有限公司")
    # res = get_log_paths(r"D:\Workspace\zhongdeng_daily\code\logs")
    # print(res)

    # df_user_outer, df_user_inner = get_user_new()
    # print("\n外部用户表:\n\n")
    # print_df(df_user_outer)
    # print("\n内部用户表:\n\n")
    # print_df(df_user_inner)

    # df_business = get_business_names_new()
    # print("\n机构信息表:\n\n")
    # print_df(df_business)
    # df_log_all = read_log_files("logs")
    # df_inner = df_log_all.merge(df_user_inner, on=STRING_COL_USER_ID)
    # res = read_log_file("logs/user_query.log")
    # res = read_log_files("logs")
    # print_df(df_inner)
    # df_inner.to_excel("debug/1.xlsx")
    pass

